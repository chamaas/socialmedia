import scrapy
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from socialmedia.items import  Posts_Insta_Item
from socialmedia.spiders import utils
from socialmedia.spiders import DRIVER_DIR , CONNEXION_DATA
from socialmedia.spiders import parameters

class InstaSpiders1_(scrapy.Spider):
    name = "InstaSpider11_"
    allowed_domains = ["wwww.instagram.com"]
    custom_settings = {
        'DOWNLOAD_DELAY': 5,
         }
    

    def __init__(self):
        """
        initialize driver
        """
        WINDOW_SIZE="1000,1000"
        self.chrome_options = Options()
        self.chrome_options.add_argument("--window-size=%s" % WINDOW_SIZE)
        #self.chrome_options.add_argument("--headless")
        #self.chrome_options.add_argument("user-agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36'")
        self.credentials = CONNEXION_DATA
        """
        pause,nb of posts,nb od query parameters
        """
        self.time_pause=3
        self.nb_postes=5
        self.index_query=[0,4]
     
    def start_requests(self):
        url= "https://www.google.com"
        yield scrapy.Request(url=url,callback=self.parse)
       
    def parse(self,response):
        url="https://www.instagram.com/?hl=fr"
        items=[]
        df = pd.DataFrame(columns=['titre','post_link','description','related_post','Likes','date_pub','url','nombre_posts_hash','nom_hash','compte','date_debut','date_fin'])
        df.to_excel('socialmedia\posts_Insta\Posts_Insta1.xlsx' , index=True, encoding='utf-8')
        self.driver = webdriver.Chrome(f"{DRIVER_DIR}",chrome_options=self.chrome_options)
        username=self.credentials['user6']["usr"]
        password= self.credentials['user6']["pass"]
        self.driver.get(url=url)
        time.sleep(2)
        
        try:
            """
            click on cookies button
            """
            cookies=self.driver.find_element_by_xpath(parameters.selector_cookies)
            cookies.click()
        except:
            pass
        time.sleep(3)
        utils.login(self.driver,username,password)
        time.sleep(5)
        try:
            """
            click on "plus tard" button
            """
            Info=self.driver.find_element_by_css_selector(parameters.selector_Info_button)
            Info.click()
        except:
            pass
        """
        return query
        """
       
        queries=utils.query()[self.index_query[0]:self.index_query[1]]
        
        if  not isinstance(queries, list):
                queries = [queries]
        date_debut=datetime.now().time()
        wb = load_workbook('socialmedia\posts_Insta\Posts_Insta1.xlsx')
        ws = wb.active
        for query in queries:
            """
            iterate n titres
            """
            try:
              titre=query.split("#")[1]
              """
              initialize item
              """
              item=Posts_Insta_Item()
              item["titre"] = titre
              if utils.get_results_hashtag(self.driver,query) is not None:
                """
                get page of hashtag
                """
                self.driver.get(url=utils.get_results_hashtag(self.driver,query))
                url_ha=self.driver.current_url
                time.sleep(1)
                nom_hash=utils.get_get_selenium_value(self.driver,parameters.selector_nom_hashtag)
                nombre_posts_hash=utils.get_get_selenium_value(self.driver,parameters.selector_nbpost)
                item["nom_hash"] = nom_hash
                item["nombre_posts_hash"] = nombre_posts_hash
                
                try:
                    """
                    click to the first post
                    """
                    Post_result=self.driver.find_element_by_css_selector(parameters.selector_firstpost)
                    Post_result.click()                            
                    time.sleep(6)
                
                    hashtags=["#livre","#livres","#books","#book","#bookstagram", "#bookstagrammer", "#bookcommunity","#booklover","#livrestagram","instalivre","#livregram","#livreaddict"]
                    try:
                       self.driver.find_element_by_xpath(parameters.selector_nextpost).click()
                    except:
                       self.driver.find_element_by_xpath("/html/body/div[6]/div[1]/div/div/div/button/div/span").click()
                          
                        
                    for i in range(self.nb_postes):
                            """
                            iterate nb posts
                            """
                            time.sleep(self.time_pause)
                            hashtags_livre=[]
                            
                            for elem in self.driver.find_elements_by_class_name('xil3i'):
                                hashtags_livre.append(elem.text)
                            """
                            verify if the post is relatad to book
                            """
                            if len(set(hashtags_livre).intersection(set(hashtags)))>0:
                                item['related_post']='related post'  
                            else:
                                item['related_post']='unrelated post'   
                            try:
                              likes=self.driver.find_element_by_css_selector(parameters.selector_likes).text
                            except:
                                pass
                           
                            date_pub = self.driver.find_element_by_css_selector(parameters.selector_datepub).get_attribute("datetime")

                            try:
                           
                              description=self.driver.find_element_by_css_selector(parameters.selector_description).text
                            except:
                              description=self.driver.find_element_by_css_selector("body > div > div > div > article > div > div > div > div > div > div > ul > div > li > div > div > div > span").text
                            """
                            extract items
                            """
                            item["url"] = url_ha 
                            item["post_link"]=self.driver.current_url
                            item["Likes"]=likes
                            item["date_pub"]=str(date_pub).split("T")[0]
                            item['description']=description.replace("\n","").replace("\t","")
                            item["compte"]=username
                            yield item
                            items.append(dict(item))
                            """
                            save item to excel
                            """
                            to_append = ["",item["titre"] ,item["post_link"],item['description'],item['related_post'],item["Likes"],item["date_pub"],item["url"],item["nombre_posts_hash"],item["nom_hash"],item["compte"]]
                            ws.append(to_append)
                            """
                            click on next post button
                            """
                            try:
                               self.driver.find_element_by_xpath(parameters.selector_nextpost2).click()
                            except:
                               self.driver.find_element_by_xpath("/html/body/div[6]/div[1]/div/div/div[2]/button/div/span").click() 
                               
                             
                except:

                            time.sleep(2)
                            self.driver.get("https://www.instagram.com/?hl=fr")
                            item=Posts_Insta_Item()
                            item["titre"] = titre
                            item["nom_hash"] = "error"
                            item["nombre_posts_hash"] = "error "
                            item['related_post']="error "
                            item["url"] = "error "
                            item["post_link"]="error "
                            item["Likes"]="error "
                            item["date_pub"]="error "
                            item['description']="error "
                            item["compte"]="error "
                            yield item
                            items.append(dict(item))
                            to_append = ["",item["titre"] ,item["nom_hash"],item["nombre_posts_hash"],item['related_post'],item["url"],item["post_link"],item["Likes"],item["date_pub"],item['description'],item["compte"]]
                            ws.append(to_append)
                            time.sleep(2)
              else:
                  item["nom_hash"] = " "
                  item["nombre_posts_hash"] = " "
                  item['related_post']=" "
                  item["url"] = " "
                  item["post_link"]=" "
                  item["Likes"]=" "
                  item["date_pub"]=" "
                  item['description']=" "
                  item["compte"]=" "
                  yield item
                  items.append(dict(item))
                  to_append = ["",item["titre"] ,item["nom_hash"],item["nombre_posts_hash"],item['related_post'],item["url"],item["post_link"],item["Likes"],item["date_pub"],item['description'],item["compte"]]
                  ws.append(to_append)
                  
            except:
                pass
        date_fin= datetime.now().time()
        ws["L2"] = date_debut
        ws["M2"] = date_fin
        wb.save('socialmedia\posts_Insta\Posts_Insta1.xlsx')
                          
     