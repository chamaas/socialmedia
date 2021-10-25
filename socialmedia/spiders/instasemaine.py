import scrapy
import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from socialmedia.items import  Insta_Item_Semaine
from socialmedia.spiders import utils
from socialmedia.spiders import DRIVER_DIR , CONNEXION_DATA
from socialmedia.spiders import parameters


class InstaSpider_semaine(scrapy.Spider):
    name = "InstaSpider_semaines"
    allowed_domains = ["wwww.instagram.com"]
    custom_settings = {
        'DOWNLOAD_DELAY': 5,
         }
    

    def __init__(self):
        """
        initialize driver
        """
        WINDOW_SIZE="1000,1000"
        self.chrome_options = Options()
        self.chrome_options.add_argument("--window-size=%s" % WINDOW_SIZE)
        #self.chrome_options.add_argument("--headless")
        #self.chrome_options.add_argument("user-agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36'")
        self.credentials = CONNEXION_DATA
        """
        pause and nb of query param
        """
        self.pause_param=6
        self.index_query=[0,4]
     
    def start_requests(self):
        url= "https://www.google.com"
        yield scrapy.Request(url=url,callback=self.parse)
       
    def parse(self,response):
        url="https://www.instagram.com/?hl=fr"
        items=[]
        self.driver = webdriver.Chrome(f"{DRIVER_DIR}",chrome_options=self.chrome_options)
        """
        create data frame
        """
        df = pd.DataFrame(columns=['Query','hashtag','nb_posts','date_debut','date_fin'])
        df.to_excel('socialmedia\hash_semaine\hashtag_Insta.xlsx' , index=True, encoding='utf-8')
        username=self.credentials['user2']["usr"]
        password= self.credentials['user2']["pass"]
        self.driver.get(url=url)
        time.sleep(2)
        
        try:
            """
            click on cookies button
            """
            cookies=self.driver.find_element_by_xpath(parameters.selector_cookies)
            cookies.click()
        except:
            pass
        time.sleep(3)
        utils.login(self.driver,username,password)
        time.sleep(5)
        try:
            """
            click on 'plus tard'
            """
            Info=self.driver.find_element_by_css_selector(parameters.selector_Info_button)
            Info.click()
        except:
            pass
        """
        return query
        """
        queries=utils.query()[self.index_query[0]:self.index_query[1]]
        date_time=datetime.now()
        
        if  not isinstance(queries, list):
                queries = [queries]
        date_debut=datetime.now().time()
        wb = load_workbook('socialmedia\hash_semaine\hashtag_Insta.xlsx')
        ws = wb.active
        for query in queries:
            """
            iterate all titre
            """
            time.sleep(self.pause_param)
            try:
              """
              initialase item
              """
              item=Insta_Item_Semaine()
              item["Query"]=query
              
              if utils.get_results_hashtag(self.driver,query) is not None:
                    nb_posts=self.driver.find_element_by_xpath(parameters.selector_nbrposts).text
                    hashtag=self.driver.find_element_by_xpath(parameters.selector_nomhash).text
                    """
                    extract items
                    """
                    item["hash"] = hashtag 
                    item["nb_posts"]=nb_posts
                    yield item
                    items.append(dict(item))
                    """
                    save item to excel
                    """
                    to_append = ["",item["Query"],item["hash"],item["nb_posts"]]
                    ws.append(to_append)
                    
              else:
                    item["hash"] = " " 
                    item["nb_posts"]=" " 
                    yield item
                    items.append(dict(item))
                    to_append = ["",item["Query"], item["hash"], item["nb_posts"]]
                    ws.append(to_append)
                    
            except:

                time.sleep(2)
                self.driver.get("https://www.instagram.com/?hl=fr")
                item=Insta_Item_Semaine()
                item["Query"]=query
                item["hash"] = "error " 
                item["nb_posts"]="error" 
                to_append = ["",item["Query"], item["hash"], item["nb_posts"]]
                ws.append(to_append)
                yield item
                time.sleep(2)
        date_fin= datetime.now().time()
        ws["E2"] = date_debut
        ws["F2"] = date_fin
        wb.save('socialmedia\hash_semaine\hashtag_Insta.xlsx')